# import libraries
from openpyxl import load_workbook
import pandas as pd
import json

# read file
wb = load_workbook("jadwal.xlsx")  # Set the filepath + filename

# select the sheet where tables are located
ws = wb["Nando"]

# check what tables that exist in the worksheet
print({key: value for key, value in ws.tables.items()})

mapping = {}

# loop through all the tables and add to a dictionary
for entry, data_boundary in ws.tables.items():
    # parse the data within the ref boundary
    data = ws[data_boundary]

    # the inner list comprehension gets the values for each cell in the table
    content = [[cell.value for cell in ent]
               for ent in data]

    header = content[0]

    rest = content[1:]

    df = pd.DataFrame(rest, columns=header)
    mapping[entry] = df.to_dict(orient="records")

json = json.dumps(mapping, indent=2)

with open("jadwal.json", "w") as out:
    out.write(json)

# # Extract all the tables to individually dataframes from the dictionary
# Tables = mapping.values()
#
# # Print each dataframe
# print(Tables)
